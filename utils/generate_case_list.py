"""
测试用例扫描工具
自动扫描 test_cases/ 目录下的所有测试文件，生成 Excel 用例列表
"""
import os
import ast
import re
import pandas as pd
from pathlib import Path
from utils.logger import logger


class TestCaseVisitor(ast.NodeVisitor):
    """
    AST 访问器，用于提取测试用例信息
    使用 NodeVisitor 模式正确跟踪当前类上下文
    """
    
    def __init__(self, module_name: str):
        self.module_name = module_name
        self.test_cases = []
        self.current_class = None  # 跟踪当前类名
        self.current_class_markers = []  # 跟踪当前类的 markers
        self.module_uri = ""  # 模块级别的 URI
    
    def visit_ClassDef(self, node):
        """
        访问类定义节点
        记录当前类名，遍历子节点后恢复
        """
        # 保存之前的类信息（支持嵌套类）
        old_class = self.current_class
        old_markers = self.current_class_markers
        
        # 设置当前类信息
        self.current_class = node.name
        self.current_class_markers = self._extract_markers(node)
        
        logger.info(f"  [DEBUG] 进入类: {self.current_class}")
        
        # 继续遍历类的子节点（方法）
        self.generic_visit(node)
        
        # 恢复之前的类信息
        self.current_class = old_class
        self.current_class_markers = old_markers
        
        logger.info(f"  [DEBUG] 离开类: {node.name}")
    
    def visit_FunctionDef(self, node):
        """
        访问函数定义节点
        只处理以 test_ 开头的测试函数
        """
        # 只处理测试函数
        if node.name.startswith("test_"):
            # 提取 Docstring
            description = self._extract_docstring(node)
            
            # 提取 API URI（优先使用函数级别，其次使用模块级别）
            api_uri = self._extract_api_uri(node)
            if not api_uri:
                api_uri = self.module_uri  # 使用模块 URI 作为默认值
            
            # 提取方法级别的 markers
            method_markers = self._extract_markers(node)
            
            # 合并类和方法的 markers
            all_markers = list(set(self.current_class_markers + method_markers))
            
            # 获取当前类名（如果在类中）
            class_name = self.current_class if self.current_class else ""
            
            logger.info(f"    [DEBUG] 找到测试函数: {node.name}, 类: {class_name}, URI: {api_uri}")
            
            # 添加到测试用例列表
            self.test_cases.append({
                "Module": self.module_name,
                "Class": class_name,
                "Function": node.name,
                "Description": description,
                "API URI": api_uri,
                "Markers": ", ".join(all_markers) if all_markers else ""
            })
        
        # 不需要继续遍历函数内部
        pass
    
    def _extract_docstring(self, node: ast.AST) -> str:
        """
        提取函数或类的 Docstring
        
        Args:
            node: AST 节点
            
        Returns:
            Docstring 内容（去除首尾空格）
        """
        docstring = ast.get_docstring(node)
        
        if docstring:
            # 去除首尾空格和多余的换行
            docstring = docstring.strip()
            
            # 提取第一行作为描述（通常是测试场景）
            lines = docstring.split("\n")
            
            # 查找 "测试场景：" 开头的行
            for line in lines:
                line = line.strip()
                if line.startswith("测试场景：") or line.startswith("测试场景:"):
                    return line.replace("测试场景：", "").replace("测试场景:", "").strip()
            
            # 如果没有找到，返回第一行
            return lines[0].strip()
        
        return ""
    
    def _extract_api_uri(self, node: ast.AST) -> str:
        """
        从 Docstring 中提取 API URI
        
        Args:
            node: AST 节点
            
        Returns:
            API URI（如 /api/v1/cores/actc/contacts）
        """
        docstring = ast.get_docstring(node)
        
        if docstring:
            # 尝试匹配多种格式的 URI
            patterns = [
                r'Path:\s*([/\w\-{}:]+)',  # Path: /api/v1/...
                r'path:\s*([/\w\-{}:]+)',  # path: /api/v1/...
                r'URI:\s*([/\w\-{}:]+)',   # URI: /api/v1/...
                r'uri:\s*([/\w\-{}:]+)',   # uri: /api/v1/...
                r'URL:\s*([/\w\-{}:]+)',   # URL: /api/v1/...
                r'url:\s*([/\w\-{}:]+)',   # url: /api/v1/...
                r'接口:\s*([/\w\-{}:]+)',  # 接口: /api/v1/...
                r'接口地址:\s*([/\w\-{}:]+)',  # 接口地址: /api/v1/...
                r'接口路径:\s*([/\w\-{}:]+)',  # 接口路径: /api/v1/...
                r'测试\s+([A-Z]+)\s+([/\w\-{}:]+)',  # 测试 GET /api/v1/...
            ]
            
            for pattern in patterns:
                match = re.search(pattern, docstring, re.IGNORECASE)
                if match:
                    # 如果是最后一个模式（包含 HTTP 方法），返回完整的 URI
                    if len(match.groups()) > 1:
                        return match.group(2)
                    else:
                        return match.group(1)
        
        return ""
    
    def _extract_markers(self, node: ast.AST) -> list:
        """
        提取 Pytest markers
        
        Args:
            node: AST 节点
            
        Returns:
            Marker 列表（如 ["contact", "list_api"]）
        """
        markers = []
        
        # 遍历装饰器
        if hasattr(node, "decorator_list"):
            for decorator in node.decorator_list:
                # 处理 @pytest.mark.xxx 形式
                if isinstance(decorator, ast.Attribute):
                    if isinstance(decorator.value, ast.Attribute):
                        # pytest.mark.xxx
                        if (hasattr(decorator.value, "attr") and 
                            decorator.value.attr == "mark" and
                            hasattr(decorator.value, "value") and
                            isinstance(decorator.value.value, ast.Name) and
                            decorator.value.value.id == "pytest"):
                            markers.append(decorator.attr)
                
                # 处理 @pytest.mark.xxx() 形式（带括号）
                elif isinstance(decorator, ast.Call):
                    if isinstance(decorator.func, ast.Attribute):
                        if isinstance(decorator.func.value, ast.Attribute):
                            if (hasattr(decorator.func.value, "attr") and 
                                decorator.func.value.attr == "mark" and
                                hasattr(decorator.func.value, "value") and
                                isinstance(decorator.func.value.value, ast.Name) and
                                decorator.func.value.value.id == "pytest"):
                                markers.append(decorator.func.attr)
        
        return markers


def extract_test_cases(test_cases_dir: str = "test_cases") -> list:
    """
    扫描测试用例目录，提取测试用例信息
    
    Args:
        test_cases_dir: 测试用例目录路径
        
    Returns:
        测试用例列表，每个元素是一个字典，包含以下字段：
        - Module: 文件名
        - Class: 测试类名
        - Function: 测试方法名
        - Description: 测试用例描述（从 Docstring 提取）
        - API URI: API 路径（从 Docstring 提取）
        - Markers: Pytest markers（如 @pytest.mark.contact）
    """
    all_test_cases = []
    
    # 获取项目根目录
    project_root = Path(__file__).parent.parent
    test_cases_path = project_root / test_cases_dir
    
    logger.info(f"[INFO] 开始扫描测试用例目录: {test_cases_path}")
    
    # 遍历所有 test_*.py 文件
    for test_file in sorted(test_cases_path.rglob("test_*.py")):
        logger.info(f"[INFO] 扫描文件: {test_file}")
        
        try:
            # 读取文件内容
            with open(test_file, "r", encoding="utf-8") as f:
                file_content = f.read()
            
            # 使用 AST 解析代码
            tree = ast.parse(file_content)
            
            # 获取相对路径作为 Module 名称
            module_name = str(test_file.relative_to(test_cases_path))
            
            # 提取模块级别的 Docstring 中的 URI
            module_docstring = ast.get_docstring(tree)
            module_uri = ""
            if module_docstring:
                # 尝试匹配 URI
                uri_patterns = [
                    r'GET\s+([/\w\-{}:]+)',  # GET /api/v1/...
                    r'POST\s+([/\w\-{}:]+)',  # POST /api/v1/...
                    r'PUT\s+([/\w\-{}:]+)',  # PUT /api/v1/...
                    r'DELETE\s+([/\w\-{}:]+)',  # DELETE /api/v1/...
                    r'Path:\s*([/\w\-{}:]+)',  # Path: /api/v1/...
                    r'测试\s+[A-Z]+\s+([/\w\-{}:]+)',  # 测试 GET /api/v1/...
                ]
                
                for pattern in uri_patterns:
                    match = re.search(pattern, module_docstring)
                    if match:
                        module_uri = match.group(1)
                        logger.info(f"  [DEBUG] 从模块 Docstring 提取 URI: {module_uri}")
                        break
            
            # 使用访问器遍历 AST
            visitor = TestCaseVisitor(module_name)
            visitor.module_uri = module_uri  # 传递模块 URI
            visitor.visit(tree)
            
            # 添加到总列表
            all_test_cases.extend(visitor.test_cases)
            
            logger.info(f"  [INFO] 找到 {len(visitor.test_cases)} 个测试用例")
        
        except Exception as e:
            logger.info(f"[ERROR] 解析文件 {test_file} 失败: {e}")
            import traceback
            traceback.print_exc()
    
    logger.info(f"[INFO] 扫描完成，共找到 {len(all_test_cases)} 个测试用例")
    
    return all_test_cases


def generate_excel(test_cases: list, output_file: str = "test_cases.xlsx"):
    """
    生成 Excel 文件
    
    Args:
        test_cases: 测试用例列表
        output_file: 输出文件名
    """
    # 创建 DataFrame
    df = pd.DataFrame(test_cases)
    
    # 获取项目根目录
    project_root = Path(__file__).parent.parent
    output_path = project_root / output_file
    
    # 生成 Excel 文件
    logger.info(f"[INFO] 生成 Excel 文件: {output_path}")
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Cases")
        
        # 获取工作表
        worksheet = writer.sheets["Test Cases"]
        
        # 设置列宽
        worksheet.column_dimensions["A"].width = 35  # Module
        worksheet.column_dimensions["B"].width = 35  # Class
        worksheet.column_dimensions["C"].width = 55  # Function
        worksheet.column_dimensions["D"].width = 80  # Description
        worksheet.column_dimensions["E"].width = 55  # API URI
        worksheet.column_dimensions["F"].width = 35  # Markers
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置内容对齐
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    logger.info(f"[INFO] Excel 文件生成成功: {output_path}")
    logger.info(f"[INFO] 共导出 {len(test_cases)} 个测试用例")


def main():
    """
    主函数
    """
    print("=" * 60)
    logger.info("测试用例扫描工具 v2.0")
    print("=" * 60)
    
    # 扫描测试用例
    test_cases = extract_test_cases()
    
    # 生成 Excel 文件
    if test_cases:
        generate_excel(test_cases)
        logger.info("\n[SUCCESS] 测试用例列表生成完成！")
    else:
        logger.info("\n[WARNING] 未找到任何测试用例")
    
    print("=" * 60)


if __name__ == "__main__":
    main()
