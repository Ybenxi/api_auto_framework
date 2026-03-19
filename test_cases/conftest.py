import pytest
import requests
import os
import sys
import json
import re
import inspect
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from api.auth_api import auth_api
from config.config import config
from utils.logger import logger
from dao.db_manager import DBManager
from utils.data_cleanup import DataCleanup

# 用于存储所有测试结果和捕获的流量
test_results = []

# 用于跟踪测试创建的ID（自动清理用）
test_created_ids = {
    "counterparty": [],
    "counterparty_group": [],
    "contact": [],
    "sub_account": [],
    "fbo_account": [],
    "financial_account": [],
    "trading_order": [],
    "client_list": [],
    "card": [],
    "user": []
}

# 模块名称映射（基于文件夹名称）
MODULE_NAME_MAPPING = {
    "contact": "Contact",
    "profile_account": "Profile Account",
    "financial_account": "Financial Account",
    "sub_account": "Sub Account",
    "fbo_account": "FBO Account",
    "statement": "Statement",
    "tenant": "Tenant",
    "open_banking": "Open Banking",
    "counterparty": "Counterparty Management",
    "identity_security": "Identity Security",
    "account_opening": "Account Opening",
    "trading_order": "Trading Order",
    "client_list": "Client List",
    "investment": "Report & Analytics",
    "account_summary": "Report & Analytics",
    "card_report": "Report & Analytics",
    "card": "Card",
    "card_opening": "Card Opening",
    "card_management": "Card Management",
    "sub_program": "Sub Program",
    "dispute_and_risk": "Dispute & Risk Control",
    "user_signup": "User Sign Up",
    "payment_deposit": "Payment & Deposit",
    "internal_pay": "Internal Pay",
    "account_transfer": "Account Transfer",
    "wire_processing": "Wire Processing",
    "remote_deposit_check": "Remote Deposit Check",
    "instant_pay": "Instant Pay",
    "ach_processing": "ACH Processing"
}


def cleanup_old_reports(reports_dir: str, keep_count: int = 5):
    """
    清理旧的带时间戳的报告文件，只保留最近的 N 次
    
    Args:
        reports_dir: 报告目录路径
        keep_count: 保留的报告数量，默认为 5
    """
    import glob
    
    try:
        # 查找所有带时间戳的报告文件
        pattern = os.path.join(reports_dir, "benxi_report_*.html")
        report_files = glob.glob(pattern)
        
        # 如果报告数量少于等于保留数量，不需要清理
        if len(report_files) <= keep_count:
            return
        
        # 按修改时间排序（最新的在前）
        report_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        # 删除多余的旧报告
        files_to_delete = report_files[keep_count:]
        deleted_count = 0
        
        for file_path in files_to_delete:
            try:
                os.remove(file_path)
                deleted_count += 1
                logger.debug(f"已删除旧报告: {os.path.basename(file_path)}")
            except Exception as e:
                logger.warning(f"删除报告失败 {os.path.basename(file_path)}: {e}")
        
        if deleted_count > 0:
            logger.info(f"报告清理完成: 删除 {deleted_count} 个旧报告，保留最近 {keep_count} 次")
    
    except Exception as e:
        logger.warning(f"清理旧报告时发生异常: {e}")


def extract_module_from_nodeid(nodeid: str) -> str:
    """
    从 nodeid 中提取模块名称
    例如: test_cases/contact/test_contact_list.py -> Contact
    """
    parts = nodeid.split("/")
    if len(parts) >= 2 and parts[0] == "test_cases":
        folder_name = parts[1]
        return MODULE_NAME_MAPPING.get(folder_name, folder_name.replace("_", " ").title())
    return "Unknown"


def extract_api_path_from_docstring(docstring: str) -> str:
    """
    从 docstring 中提取 API 路径
    支持格式：
    - 测试 GET /api/v1/cores/{core}/contacts 接口
    - Path: /api/v1/...
    - URI: /api/v1/...
    """
    if not docstring:
        return ""
    
    # 匹配模式1: "测试 GET/POST/PUT/DELETE /api/..." 或 "GET /api/..."
    pattern1 = r'(?:测试\s+)?(?:GET|POST|PUT|DELETE|PATCH)\s+(/api/[^\s]+)'
    match = re.search(pattern1, docstring, re.IGNORECASE)
    if match:
        return match.group(1)
    
    # 匹配模式2: "Path: /..." 或 "URI: /..."
    pattern2 = r'(?:Path|URI|Endpoint):\s*(/[^\s]+)'
    match = re.search(pattern2, docstring, re.IGNORECASE)
    if match:
        return match.group(1)
    
    # 匹配模式3: 直接匹配 /api/v1/... 路径
    pattern3 = r'(/api/v\d+/[^\s]+)'
    match = re.search(pattern3, docstring)
    if match:
        return match.group(1)
    
    return ""


def extract_docstring_summary(item) -> str:
    """
    提取测试用例的 docstring 摘要（第一行非空内容，即中文描述）
    """
    try:
        func = item.obj if hasattr(item, 'obj') else None
        if func and func.__doc__:
            lines = func.__doc__.strip().split('\n')
            for line in lines:
                line = line.strip()
                if line:
                    return line
    except Exception:
        pass
    return ""


def extract_docstring_en(item) -> str:
    """
    提取 docstring 的第二行非空内容作为英文描述。
    双行规范：
      第一行：中文描述（如 "测试场景1：成功获取订单列表"）
      第二行：英文描述（如 "Test Scenario1: Successfully Retrieve Order List"）
    如果没有第二行，fallback 到测试函数名（已是英文命名规范）。
    """
    try:
        func = item.obj if hasattr(item, 'obj') else None
        if func and func.__doc__:
            lines = [l.strip() for l in func.__doc__.strip().split('\n') if l.strip()]
            # lines[0] = 中文，lines[1] = 英文（如果存在）
            if len(lines) >= 2:
                return lines[1]
    except Exception:
        pass
    return ""


def get_full_docstring(item) -> str:
    """
    获取测试用例的完整 docstring
    """
    try:
        func = item.obj if hasattr(item, 'obj') else None
        if func and func.__doc__:
            return func.__doc__.strip()
        
        # 尝试从类级别获取 docstring
        if hasattr(item, 'cls') and item.cls and item.cls.__doc__:
            return item.cls.__doc__.strip()
    except Exception:
        pass
    return ""


def get_module_docstring(item) -> str:
    """
    获取测试模块（文件）的 docstring
    """
    try:
        if hasattr(item, 'module') and item.module and item.module.__doc__:
            return item.module.__doc__.strip()
    except Exception:
        pass
    return ""


def extract_scenario_number(docstring_summary: str) -> int:
    """
    从 docstring 摘要中提取测试场景编号
    例如: "测试场景8：排序功能验证" -> 8
    如果没有编号，返回 999（排到最后）
    """
    if not docstring_summary:
        return 999
    
    # 匹配模式：测试场景X：或 Test ScenarioX:
    pattern = r'(?:测试场景|Test Scenario)\s*(\d+)'
    match = re.search(pattern, docstring_summary)
    
    if match:
        return int(match.group(1))
    
    return 999  # 没有编号的排到最后


def translate_docstring_to_english(zh_text: str) -> str:
    """
    将中文测试场景描述转换为英文
    使用模式匹配和映射表
    """
    if not zh_text:
        return ""
    
    # 中英文映射表（完整版 v2.0 - 扩充50+词汇）
    translations = {
        # 完整短语（优先匹配）
        "排序功能验证 - 按姓名排序": "Sorting Verification - Sort by Name",
        "排序功能验证 - 按账户名称排序": "Sorting Verification - Sort by Account Name",
        "创建后立即查询详情，验证数据一致性": "Create then Query Detail Immediately to Verify Data Consistency",
        "创建后立即在列表中查询，验证数据一致性": "Create then Query in List Immediately to Verify Data Consistency",
        "对比详情接口和列表接口返回的数据一致性": "Compare Data Consistency between Detail and List API",
        "使用不存在的筛选条件，验证空结果处理": "Verify Empty Result with Non-existent Filter",
        "响应字段完整性验证": "Response Fields Completeness Verification",
        "字段完整性验证": "Fields Completeness Verification",
        
        # 测试场景标识
        "测试场景": "Test Scenario",
        
        # 操作类型（完整短语）
        "基础列表查询 - 验证接口可用性": "Basic List Query - Verify API Availability",
        "排序功能验证": "Sorting Verification",
        "分页功能验证": "Pagination Verification",
        "空结果验证": "Empty Result Verification",
        "空字符串边界值测试": "Empty String Boundary Test",
        
        # CRUD 操作
        "基础列表查询": "Basic List Query",
        "验证接口可用性": "Verify API Availability",
        "名称筛选查询 - 验证筛选逻辑": "Name Filter - Verify Filter Logic",
        "名称筛选查询": "Name Filter Query",
        "名称筛选": "Name Filter",
        "枚举类型筛选": "Enum Type Filter",
        "状态筛选": "Status Filter",
        "多条件组合筛选": "Multiple Filters Combined",
        
        "成功获取": "Successfully Retrieve",
        "成功创建": "Successfully Create",
        "成功更新": "Successfully Update",
        "使用必需字段创建": "Create with Required Fields",
        "使用所有字段创建": "Create with All Fields",
        "创建后立即查询详情": "Create then Query Detail Immediately",
        "创建后立即在列表中查询": "Create then Query in List Immediately",
        "创建后立即查询": "Query Immediately After Creation",
        "缺少必需字段时创建失败": "Create Failed with Missing Required Fields",
        "缺少必需字段": "Missing Required Fields",
        "使用无效": "With Invalid",
        "使用不存在": "With Non-existent",
        "同时更新多个字段": "Update Multiple Fields",
        
        # 特定描述
        "包括可选字段": "Including Optional Fields",
        "包含加密": "With Encrypted",
        "并包含加密": "With Encrypted",
        "无效 email 格式": "Invalid Email Format",
        "无效 phone 格式": "Invalid Phone Format",
        "非 E.164 格式": "Non-E.164 Format",
        "验证筛选逻辑": "Verify Filter Logic",
        "验证所有字段": "Verify All Fields",
        
        # 资源类型和名词（大幅扩充）
        "账户": "Account",
        "联系人": "Contact",
        "详情": "Detail",
        "列表": "List",
        "关联": "Related",
        "地址": "Address",
        "邮寄地址": "Mailing Address",
        "注册地址": "Register Address",
        "永久地址": "Permanent Address",
        "部分字段": "Partial Fields",
        "姓名": "Name",
        "账户名称": "Account Name",
        "用户": "User",
        "个人资料": "Profile",
        "资料": "Profile",  # 新增
        "头像": "Avatar",
        "密码": "Password",
        "因子": "Factor",
        "验证码": "Pass Code",
        "凭证": "Credentials",
        "文件": "File",
        "参数": "Parameter",  # 新增
        "类型": "Type",  # 新增
        "数据": "Data",  # 新增
        "信息": "Information",  # 新增
        "方式": "Method",  # 新增
        "操作": "Operation",  # 新增
        "请求": "Request",  # 新增
        "返回": "Return",  # 新增
        "结果": "Result",  # 新增
        "内容": "Content",  # 新增
        "记录": "Record",  # 新增
        "条件": "Condition",  # 新增
        "范围": "Range",  # 新增
        "限制": "Limit",  # 新增
        "权限": "Permission",  # 新增
        "配置": "Configuration",  # 新增
        "设置": "Setting",  # 新增
        "选项": "Option",  # 新增
        "状态码": "Status Code",  # 新增
        "错误码": "Error Code",  # 新增
        "交易": "Transaction",  # 新增
        "支付": "Payment",  # 新增
        "账号": "Account Number",  # 新增
        "金额": "Amount",  # 新增
        "余额": "Balance",  # 新增
        "日期": "Date",  # 新增
        "时间": "Time",  # 新增
        "期限": "Duration",  # 新增
        "频率": "Frequency",  # 新增
        "完整性": "Completeness",  # 新增
        "对手方": "Counterparty",  # 新增
        "交易对手": "Counterparty",  # 新增
        "分组": "Group",  # 新增
        "成员": "Member",  # 新增
        "身份": "Identity",  # 新增
        "安全": "Security",  # 新增
        "认证": "Authentication",  # 新增
        "授权": "Authorization",  # 新增
        
        # 验证相关
        "验证点": "Verification Points",
        "验证数据一致性": "Verify Data Consistency",
        "数据一致性": "Data Consistency",
        "对比": "Compare",
        "响应": "Response",
        "结构": "Structure",
        "数据结构": "Data Structure",
        
        # 动词（扩充）
        "更新": "Update",
        "创建": "Create",
        "筛选": "Filter",
        "查询": "Query",
        "获取": "Retrieve",
        "验证": "Verify",
        "测试": "Test",
        "排序": "Sort",
        "删除": "Delete",  # 新增
        "添加": "Add",  # 新增
        "移除": "Remove",  # 新增
        "发送": "Send",  # 新增
        "接收": "Receive",  # 新增
        "上传": "Upload",  # 新增
        "下载": "Download",  # 新增
        "修改": "Modify",  # 新增
        "设置": "Set",  # 新增
        "清空": "Clear",  # 新增
        "重置": "Reset",  # 新增
        "刷新": "Refresh",  # 新增
        "同步": "Sync",  # 新增
        "导出": "Export",  # 新增
        "导入": "Import",  # 新增
        "生成": "Generate",  # 新增
        "计算": "Calculate",  # 新增
        "处理": "Process",  # 新增
        "执行": "Execute",  # 新增
        "调用": "Call",  # 新增
        "触发": "Trigger",  # 新增
        "缺少": "Missing",  # 新增
        
        # 形容词和状态（扩充）
        "成功": "Success",
        "失败": "Failed",
        "无效": "Invalid",
        "有效": "Valid",
        "完整": "Complete",
        "所有": "All",
        "全部": "All",
        "多个": "Multiple",
        "单个": "Single",
        "必需": "Required",
        "可选": "Optional",
        "正确": "Correct",  # 新增
        "错误": "Error",  # 新增
        "空": "Empty",  # 新增
        "非空": "Non-empty",  # 新增
        "重复": "Duplicate",  # 新增
        "唯一": "Unique",  # 新增
        "有效": "Valid",
        "已": "",  # 新增（去除）
        "未": "Not ",  # 新增
        
        # 连词和介词（扩充）
        "并": " and ",
        "与": " and ",
        "和": " and ",  # 新增
        "或": " or ",
        "或者": " or ",  # 新增
        "但": " but ",  # 新增
        "但是": " but ",  # 新增
        "如果": " if ",  # 新增
        "当": " when ",  # 新增
        "在": " in ",  # 新增
        "对": " for ",  # 新增
        "到": " to ",  # 新增
        "从": " from ",  # 新增
        "于": " in ",  # 新增
        "向": " to ",  # 新增
        "给": " to ",  # 新增
        "按": "by ",
        "使用": "with ",
        "通过": "via ",  # 新增
        "根据": "according to ",  # 新增
        "基于": "based on ",  # 新增
        "关于": "about ",  # 新增
        
        # 其他词汇（扩充）
        "包含": "Contains",
        "不包含": "Not Contains",
        "边界值": "Boundary Value",
        "格式": "Format",
        "错误": "Error",
        "字段": "Fields",
        "加密": "Encrypted",
        "接口": "API",
        "功能": "Function",
        "场景": "Scenario",  # 新增
        "情况": "Case",  # 新增
        "步骤": "Step",  # 新增
        "流程": "Flow",  # 新增
        "过程": "Process",  # 新增
        "方法": "Method",  # 新增
        "函数": "Function",  # 新增
        "模块": "Module",  # 新增
        "组件": "Component",  # 新增
        "服务": "Service",  # 新增
        "系统": "System",  # 新增
        "应用": "Application",  # 新增
        "程序": "Program",  # 新增
        "代码": "Code",  # 新增
        "脚本": "Script",  # 新增
        "工具": "Tool",  # 新增
        "辅助": "Helper",  # 新增
        "实用": "Utility",  # 新增
        "公共": "Common",  # 新增
        "私有": "Private",  # 新增
        "保护": "Protected",  # 新增
        "静态": "Static",  # 新增
        "动态": "Dynamic",  # 新增
        "临时": "Temporary",  # 新增
        "永久": "Permanent",  # 新增
        
        # 语气词和连接符（删除）
        "的": " ",
        "了": "",
        " - ": " - ",
        "时": " when",
        "性": "",  # 新增（去除-ness后缀）
        "化": "",  # 新增（去除-ization后缀）
        
        # 标点符号
        "：": ": ",
        "，": ", ",
        "（": " (",
        "）": ")",
        "、": " & "
    }
    
    result = zh_text
    # 按照映射表长度降序排序，优先匹配长字符串
    for zh, en in sorted(translations.items(), key=lambda x: len(x[0]), reverse=True):
        result = result.replace(zh, en)
    
    # 清理多余空格
    result = re.sub(r'\s+', ' ', result).strip()
    
    return result


@pytest.fixture(scope="session", autouse=True)
def login_session():
    """
    全局登录会话管理
    """
    session = requests.Session()
    # 强制要求 JSON
    session.headers.update({"Content-Type": "application/json; charset=utf-8"})
    
    # 执行登录逻辑
    logger.info("正在执行全局登录...")
    try:
        # 1. 获取认证配置
        auth_data = config.auth_data
        url = f"{config.base_url}/api/v1/auth/{auth_data['tenant_id']}/oauth2/token"
        params = {
            "grant_type": "client_credentials",
            "user_id": auth_data["user_id"],
            "Jmeter-Test": "Jmeter-Test"
        }
        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Basic {auth_data['basic_auth']}"
        }
        
        # 2. 发送登录请求 (注意：登录请求不使用 session，避免 header 冲突)
        response = requests.post(url, params=params, headers=headers)
        
        if response.status_code == 200:
            res_json = response.json()
            # 兼容性处理：Token 可能在根目录，也可能在 data 目录下
            token = res_json.get("access_token") or res_json.get("data", {}).get("access_token")
            if token:
                session.headers.update({"Authorization": f"Bearer {token}"})
                logger.info(f"登录成功，获取 Token: {token[:10]}...")
            else:
                logger.warning(f"登录响应中未找到 access_token, 响应内容: {res_json}")
        else:
            logger.error(f"登录失败，状态码: {response.status_code}, 响应: {response.text}")
    except Exception as e:
        logger.error(f"登录过程发生异常: {e}", exc_info=True)

    yield session
    
    logger.info("正在关闭会话...")
    session.close()


@pytest.fixture(scope="session")
def db_session():
    """
    数据库会话管理（可选）
    仅在需要数据库验证时使用
    """
    db_manager = None
    try:
        # 尝试初始化数据库连接
        db_config = config.db_config
        db_manager = DBManager(db_config)
        logger.info("数据库会话已就绪")
        yield db_manager
    except Exception as e:
        logger.warning(f"数据库会话初始化跳过: {e}")
        yield None
    finally:
        if db_manager:
            db_manager.close_connection()
            logger.info("数据库会话已关闭")


@pytest.fixture(scope="session")
def db_cleanup():
    """
    数据库清理器 fixture（session级别）
    用于自动清理测试过程中创建的垃圾数据
    
    ⚠️ 重要：
    1. 测试结束后自动清理跟踪的ID
    2. 只删除名称以 "Auto TestYan" 开头的数据
    3. 默认先模拟，确认无误后再实际删除
    
    使用方法：
        def test_create_something(api, db_cleanup):
            response = api.create(...)
            created_id = response.json()["data"]["id"]
            track_created_id("counterparty", created_id)  # 自动清理
    """
    db_manager = None
    cleaner = None
    
    try:
        # 初始化数据库连接
        db_config = config.db_config
        
        # 检查是否配置了数据库（避免使用默认的localhost）
        if db_config.get("host") == "localhost":
            logger.warning("数据库未配置（使用默认localhost），跳过数据清理功能")
            yield None
            return
        
        db_manager = DBManager(db_config)
        cleaner = DataCleanup(db_manager)
        
        logger.info("=" * 60)
        logger.info("数据清理器已就绪")
        logger.info("⚠️  测试结束后将自动清理跟踪的测试数据")
        logger.info("=" * 60)
        
        yield cleaner
        
        # 测试结束后清理
        logger.info("")
        logger.info("=" * 60)
        logger.info("测试会话结束，开始清理数据...")
        logger.info("=" * 60)
        
        total_cleaned = 0
        # 优先使用 cleaner 实例上跟踪的 ID（避免模块级变量多实例问题）
        tracked = cleaner._tracked_ids if cleaner else {}
        for module, ids in tracked.items():
            if ids:
                logger.info(f"\n清理 {module}: {len(ids)} 条记录")
                logger.debug(f"ID列表: {ids}")
                
                try:
                    # 实际执行清理（dry_run=False）
                    stats = cleaner.cleanup_by_ids(module, ids, dry_run=False)
                    
                    # 统计总删除数
                    module_total = sum(stats.values())
                    total_cleaned += module_total
                    
                    logger.info(f"✓ {module} 清理完成: {stats}")
                    
                except Exception as e:
                    logger.error(f"✗ {module} 清理失败: {e}", exc_info=True)
        
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"数据清理完成！共清理 {total_cleaned} 条记录")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"数据库清理器初始化失败: {e}", exc_info=True)
        yield None
    
    finally:
        if db_manager:
            db_manager.close_connection()


def track_created_id(module: str, resource_id: str):
    """
    跟踪测试创建的ID（用于测试结束后自动清理）
    
    Args:
        module: 模块名称（如 "counterparty", "contact", "sub_account"）
        resource_id: 资源ID
    
    Example:
        # 在测试中使用
        response = counterparty_api.create_counterparty(...)
        created_id = response.json()["data"]["id"]
        track_created_id("counterparty", created_id)
        
        # 测试结束后会自动清理
    """
    if module in test_created_ids:
        test_created_ids[module].append(str(resource_id))
        logger.debug(f"✓ 已跟踪 {module} ID: {resource_id}")
    else:
        logger.warning(f"未知模块: {module}，无法跟踪ID")


def get_tracked_ids(module: str) -> list:
    """
    获取已跟踪的ID列表
    
    Args:
        module: 模块名称
    
    Returns:
        ID列表
    """
    return test_created_ids.get(module, [])


def _analyze_failure(longrepr: str, extra_data: dict) -> str:
    """
    根据 pytest 的 longrepr（失败信息）和接口流量数据，自动分析失败原因，
    返回简明扼要的中文分析结论。
    """
    text = longrepr or ""
    lines = text.splitlines()

    # ── 提取 AssertionError 信息 ──────────────────────────────────
    assert_lines = [l.strip() for l in lines if l.strip().startswith("AssertionError")]
    assert_msg = assert_lines[0] if assert_lines else ""

    # ── 提取最后一条 assert 语句（E  assert ... 行）────────────────
    e_lines = [l.strip() for l in lines if l.strip().startswith("E ") or l.strip().startswith("E\t")]
    e_msg = " | ".join(e_lines[:5]) if e_lines else ""

    # ── 提取接口响应信息 ──────────────────────────────────────────
    resp = {}
    if extra_data and extra_data.get("response"):
        resp = extra_data["response"].get("body", {}) or {}
    resp_code = resp.get("code") if isinstance(resp, dict) else None
    resp_err  = resp.get("error_message") if isinstance(resp, dict) else None
    http_code = extra_data.get("response", {}).get("status_code") if extra_data else None

    # ── 规则匹配，输出简明结论 ─────────────────────────────────────

    # 1. pytest.skip 被触发
    if "Skipped" in text or "skip" in text.lower() and "pytest.skip" in text:
        return "测试被 pytest.skip 跳过，不计入失败"

    # 2. HTTP 连接 / 超时
    if any(k in text for k in ["ConnectionError", "ConnectTimeout", "ReadTimeout", "MaxRetryError"]):
        return "网络连接失败或请求超时，可能是环境不可达或网络不稳定"

    # 3. 接口返回非预期业务 code
    if resp_code is not None and resp_code != 200:
        err_hint = f"，错误信息：「{resp_err}」" if resp_err else ""
        return f"API 业务层返回错误 code={resp_code}{err_hint}"

    # 4. HTTP 状态码非 200
    if http_code and http_code != 200:
        return f"HTTP 请求失败，状态码 {http_code}（期望 200）"

    # 5. 字段缺失断言
    if "缺少必需字段" in text or "missing" in text.lower() or "not in" in e_msg.lower():
        field_match = re.search(r"['\"]([a-z_]+)['\"]", e_msg or text)
        field_hint = f"（字段：{field_match.group(1)}）" if field_match else ""
        return f"响应中缺少预期字段{field_hint}"

    # 6. 值不匹配（期望 vs 实际）
    if "期望" in text and "实际" in text:
        # 优先匹配「期望 XXX，实际 YYY」或「期望「XXX」，实际「YYY」」两种常见格式
        m = re.search(r"期望\s*「([^」]{1,80})」[，,]?\s*实际\s*「([^」]{1,80})」", text)
        if not m:
            m = re.search(r"期望\s+'([^']{1,80})'\s*[，,]?\s*实际\s+'([^']{1,80})'", text)
        if not m:
            # 无引号：期望后取到逗号/换行，实际后取到逗号/换行/句末
            m = re.search(r"期望\s+([^\n，,。]{1,60})[，,]\s*实际\s+([^\n，,。]{1,60})", text)
        if m:
            return f"字段值不匹配：期望「{m.group(1).strip()}」，实际「{m.group(2).strip()}」"
        return "字段值与预期不一致"

    # 7. assert False / assert x == y
    if e_msg:
        # 截取关键部分，避免太长
        short = e_msg[:200].replace("\n", " ").strip()
        return f"断言失败：{short}"

    # 8. AssertionError 兜底
    if assert_msg:
        return assert_msg[:200]

    # 9. 异常类型兜底
    exc_match = re.search(r"([\w.]+Error|[\w.]+Exception)[:\s]+(.*)", text)
    if exc_match:
        return f"{exc_match.group(1)}：{exc_match.group(2)[:150].strip()}"

    # 10. 完全兜底
    last_lines = [l.strip() for l in lines if l.strip()]
    return last_lines[-1][:200] if last_lines else "未知失败原因，请查看失败详情"


# --- 流量捕获逻辑（增强版 - 防止重复记录）---
@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    
    if report.when == 'call':
        # 获取用例中存储的流量数据
        extra_data = getattr(item, "extra_data", {})
        
        # 提取 marker 信息
        markers = [marker.name for marker in item.iter_markers()]
        
        # 提取模块名称（从文件夹名称映射）
        module_name = extract_module_from_nodeid(item.nodeid)
        
        # 提取测试文件名（不包含路径）
        nodeid_parts = item.nodeid.split("::")
        test_file = nodeid_parts[0].split("/")[-1].replace(".py", "")
        
        # 提取 docstring 信息
        docstring_summary_zh = extract_docstring_summary(item)
        full_docstring = get_full_docstring(item)
        module_docstring = get_module_docstring(item)
        
        # 生成英文版本的 docstring（优先读第二行，fallback 到函数名）
        docstring_en_from_line2 = extract_docstring_en(item)
        if docstring_en_from_line2:
            docstring_summary_en = docstring_en_from_line2
        else:
            # 没有第二行时 fallback：用函数名（已是英文），比词典替换更干净
            test_func_name_en = item.nodeid.split("::")[-1]
            docstring_summary_en = test_func_name_en
        
        # 提取 API 路径（优先从方法 docstring，其次从模块 docstring）
        api_path = extract_api_path_from_docstring(full_docstring)
        if not api_path:
            api_path = extract_api_path_from_docstring(module_docstring)
        
        # 提取测试函数名（用于显示）
        test_func_name = item.nodeid.split("::")[-1]
        
        # 提取场景编号（用于排序）
        scenario_number = extract_scenario_number(docstring_summary_zh)

        # 提取失败原因和智能分析（仅失败用例）
        failure_reason = None
        failure_analysis = None
        if report.outcome == "failed" and report.longrepr:
            failure_reason = str(report.longrepr)
            failure_analysis = _analyze_failure(failure_reason, extra_data)

        # 构建测试结果数据
        test_data = {
            "nodeid": item.nodeid,
            "status": report.outcome,
            "duration": report.duration,
            "start_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "markers": markers,
            "module": module_name,
            "test_file": test_file,
            "test_func_name": test_func_name,
            "docstring_summary_zh": docstring_summary_zh,
            "docstring_summary_en": docstring_summary_en,
            "docstring_summary": docstring_summary_zh,  # 保持向后兼容
            "scenario_number": scenario_number,  # 用于排序
            "api_path": api_path,
            "failure_reason": failure_reason,
            "failure_analysis": failure_analysis,
            "extra": extra_data
        }
        
        # 防止重复记录（处理 pytest-rerunfailures 重试）
        # 查找是否已有该 nodeid 的记录
        existing_index = None
        for i, result in enumerate(test_results):
            if result["nodeid"] == item.nodeid:
                existing_index = i
                break
        
        if existing_index is not None:
            # 如果已存在，更新为最新结果（重试后的最终结果）
            test_results[existing_index] = test_data
        else:
            # 如果不存在，追加新记录
            test_results.append(test_data)


def _generate_excel_from_results(results: list, output_path: str):
    """
    从 test_results（与 HTML 报告完全相同的数据源）生成 Excel。
    规则：
      - 所有文字纯英文（字段名、状态、标记等）
      - 包含请求/响应信息（method, url, request_body, status_code, response_body）
      - 按模块（module 字段）分 sheet，每个 sheet 含该模块所有用例
      - 额外保留一个 "All" sheet 汇总所有数据
    """
    import pandas as pd
    from pathlib import Path
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def _safe_str(v, max_len=32767):
        """Excel 单元格最大 32767 字符"""
        s = str(v) if v is not None else ""
        return s[:max_len] if len(s) > max_len else s

    def _flatten(item: dict) -> dict:
        """将一条 test_result 展开为 Excel 一行（纯英文字段名）"""
        extra = item.get("extra") or {}
        req = extra.get("request") or {}
        resp = extra.get("response") or {}
        import json as _json
        req_body = req.get("body")
        resp_body = resp.get("body")
        return {
            "Module":            _safe_str(item.get("module", "")),
            "Test File":         _safe_str(item.get("test_file", "")),
            "Test Case (EN)":    _safe_str(item.get("docstring_summary_en") or item.get("test_func_name", "")),
            "API Path":          _safe_str(item.get("api_path", "")),
            "Status":            _safe_str((item.get("status") or "").upper()),
            "Duration (s)":      round(float(item.get("duration") or 0), 3),
            "Start Time":        _safe_str(item.get("start_time", "")),
            "Request Method":    _safe_str(req.get("method", "")),
            "Request URL":       _safe_str(req.get("url", "")),
            "Request Body":      _safe_str(_json.dumps(req_body, ensure_ascii=False) if req_body else ""),
            "HTTP Status Code":  _safe_str(resp.get("status_code", "")),
            "Response Body":     _safe_str(_json.dumps(resp_body, ensure_ascii=False) if resp_body else ""),
            "Failure Analysis":  _safe_str(item.get("failure_analysis") or ""),
            "Failure Detail":    _safe_str(item.get("failure_reason") or "", max_len=5000),
        }

    COLUMN_WIDTHS = {
        "Module": 22, "Test File": 35, "Test Case (EN)": 60, "API Path": 45,
        "Status": 10, "Duration (s)": 12, "Start Time": 20,
        "Request Method": 12, "Request URL": 70, "Request Body": 50,
        "HTTP Status Code": 14, "Response Body": 60,
        "Failure Analysis": 50, "Failure Detail": 60,
    }
    STATUS_COLORS = {
        "PASSED":  "C6EFCE", "FAILED":  "FFC7CE", "SKIPPED": "FFEB9C",
        "ERROR":   "FFC7CE",
    }

    def _style_sheet(ws, df):
        # 表头
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill(start_color="2F54EB", end_color="2F54EB", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 20)
        ws.row_dimensions[1].height = 30
        # 数据行：Status 列着色
        status_col = None
        for i, col_name in enumerate(df.columns, 1):
            if col_name == "Status":
                status_col = i
                break
        for row_idx in range(2, ws.max_row + 1):
            status_val = ws.cell(row=row_idx, column=status_col).value if status_col else ""
            bg = STATUS_COLORS.get(str(status_val).upper(), "FFFFFF")
            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    rows = [_flatten(item) for item in results]
    if not rows:
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # 按模块分组
    from collections import defaultdict
    modules = defaultdict(list)
    for row in rows:
        modules[row["Module"]].append(row)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 每个模块一个 sheet
        for module_name, module_rows in sorted(modules.items()):
            sheet_name = (module_name or "Unknown")[:31]  # Excel sheet 名限 31 字符
            df = pd.DataFrame(module_rows)
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            _style_sheet(writer.sheets[sheet_name], df)

        # 汇总 All sheet
        df_all = pd.DataFrame(rows)
        df_all.to_excel(writer, index=False, sheet_name="All")
        _style_sheet(writer.sheets["All"], df_all)


def pytest_sessionfinish(session, exitstatus):
    """
    测试结束时，将结果注入 HTML 模板
    使用时间戳命名报告文件，避免覆盖
    """
    template_path = os.path.join(os.path.dirname(__file__), "..", "assets", "report_template.html")
    
    # 生成带时间戳的报告文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"benxi_report_{timestamp}.html"
    report_path = os.path.join(os.path.dirname(__file__), "..", "reports", report_filename)
    
    if not os.path.exists(os.path.dirname(report_path)):
        os.makedirs(os.path.dirname(report_path))
    
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
        
        # 三级排序：模块 -> 测试文件 -> 场景编号
        sorted_results = sorted(
            test_results, 
            key=lambda x: (
                x.get("module", ""),         # 第1级：模块名称
                x.get("test_file", ""),      # 第2级：测试文件名
                x.get("scenario_number", 999),  # 第3级：场景编号（从 docstring 提取）
                x.get("nodeid", "")          # 第4级：nodeid（兜底）
            )
        )
        
        # 注入数据
        # 注意：json.dumps 后需要将 </ 转义为 <\/，防止响应体中的 </script> 等 HTML 标签
        # 破坏 <script> 块结构（RFC 4627 / HTML5 规范允许此转义）
        env_info = {
            "env": os.getenv("ENV", "DEV"),
            "core": config.core
        }
        results_json = json.dumps(sorted_results, ensure_ascii=False).replace("</", "<\\/")
        env_json = json.dumps(env_info, ensure_ascii=False).replace("</", "<\\/")
        final_html = template_content.replace("{{RESULTS_JSON}}", results_json)
        final_html = final_html.replace("{{ENV_JSON}}", env_json)
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
        logger.info(f"Ben xi report v3.0 已生成: {report_path}")
        
        # 同时生成一个固定名称的报告（用于 CI/CD）
        latest_report_path = os.path.join(os.path.dirname(__file__), "..", "reports", "final_report.html")
        with open(latest_report_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
        logger.info(f"最新报告已更新: {latest_report_path}")
        
        # 报告永久保留，不再自动清理（可在测试平台手动删除）

        # 自动生成 Excel 测试用例清单（带时间戳，与 HTML 报告命名一致）
        excel_filename = f"test_cases_{timestamp}.xlsx"
        excel_path = os.path.join(os.path.dirname(__file__), "..", "reports", excel_filename)
        try:
            _generate_excel_from_results(sorted_results, excel_path)
            logger.info(f"Excel 测试用例清单已生成: reports/{excel_filename}")
        except Exception as e:
            logger.warning(f"Excel 测试用例清单生成失败（不影响报告）: {e}")

        # 自动生成 PDF 摘要报告（管理层用）
        pdf_filename = f"summary_report_{timestamp}.pdf"
        pdf_path = os.path.join(os.path.dirname(__file__), "..", "reports", pdf_filename)
        try:
            sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
            from utils.generate_pdf_report import generate_pdf_summary
            env_name = os.getenv("ENV", "DEV")
            generate_pdf_summary(sorted_results, pdf_path, env=env_name, core=config.core)
            logger.info(f"PDF 摘要报告已生成: reports/{pdf_filename}")
            # 同时生成一个固定名称（供平台直接引用）
            pdf_latest = os.path.join(os.path.dirname(__file__), "..", "reports", "final_summary.pdf")
            import shutil
            shutil.copy2(pdf_path, pdf_latest)
        except Exception as e:
            logger.warning(f"PDF 摘要报告生成失败（不影响测试）: {e}")


# --- 自动拦截 Requests 流量的辅助工具（修复版 v2）---
@pytest.fixture(autouse=True)
def capture_traffic(request):
    """
    自动拦截用例中的所有 requests.Session 调用并记录。
    v2：改为 patch requests.Session.request（全局），
    解决 identity_security 等模块使用独立 session 时流量无法捕获的问题。
    """
    import requests as _requests

    original_method = _requests.Session.request

    def patched_request(self_session, method, url, **kwargs):
        response = original_method(self_session, method, url, **kwargs)
        try:
            # 解析响应 body
            try:
                res_body = response.json()
            except Exception:
                res_body = response.text

            # 构建请求参数
            req_body = {}
            if method.upper() == "GET":
                parsed_url = urlparse(url)
                query_params = parse_qs(parsed_url.query)
                req_body = {k: v[0] if len(v) == 1 else v for k, v in query_params.items()}
                if 'params' in kwargs and kwargs['params']:
                    req_body.update(kwargs['params'])
            else:
                if 'json' in kwargs and kwargs['json']:
                    req_body = kwargs['json']
                elif 'data' in kwargs and kwargs['data']:
                    req_body = kwargs['data']

            traffic = {
                "request": {"method": method.upper(), "url": url, "body": req_body},
                "response": {"status_code": response.status_code, "body": res_body}
            }
            request.node.extra_data = traffic
        except Exception as e:
            logger.warning(f"流量捕获失败: {e}")
            if not hasattr(request.node, 'extra_data'):
                request.node.extra_data = {}

        return response

    _requests.Session.request = patched_request
    yield
    _requests.Session.request = original_method
